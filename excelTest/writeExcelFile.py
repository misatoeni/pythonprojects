##read a json file and write a excel file using the data

import json
from openpyxl import Workbook
from openpyxl.styles import Alignment

JSON_FILE = 'poblacionHistoricoMexico.json'
EXCEL_FILE = 'Poblacion_Mexico_1960-2020.xlsx'

##read the json file
with open(JSON_FILE, 'r') as openfile:
    datos = json.load(openfile)

##create the workbook
workbook = Workbook()

##active worksheet
worksheet = workbook.active

worksheet['A1'].value = 'Población México 1960-2020'
worksheet['A1'].alignment = Alignment(horizontal='center',vertical='center')
worksheet.merge_cells('A1:A61')


datosPoblacion = datos[1]
count = 1
for dato in datosPoblacion:
    if dato['value'] != None:
        worksheet.cell(row = count, column = 2, value = dato['date'])
        worksheet.cell(row = count, column = 3, value = dato['value'])
        count = count + 1


##save the file
workbook.save(EXCEL_FILE)

